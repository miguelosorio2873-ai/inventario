import customtkinter as ctk
from config import *
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import filedialog, messagebox
from collections import OrderedDict
from utils.icon_util import icon_chart_pie, icon_chart_bar, icon_download
import os
import subprocess

from dao.producto_dao import ProductoDAO
from dao.categoria_dao import CategoriaDAO
from dao.cliente_dao import ClienteDAO
from dao.proveedor_dao import ProveedorDAO
from dao.inventario_dao import InventarioDAO
from dao.factura_dao import FacturaDAO
from dao.usuario_dao import UsuarioDAO


SDF_FMT = "%Y-%m-%d"

COLUMNAS_POR_TIPO = OrderedDict([
    ("Productos", ["ID", "SKU", "Nombre", "Descripción", "Categoría", "Precio Venta", "Costo Promedio", "Stock Mínimo", "Stock Actual", "Estado"]),
    ("Categorías", ["ID", "Nombre", "Descripción"]),
    ("Clientes", ["ID", "Cédula", "Nombre", "Correo", "Teléfono"]),
    ("Proveedores", ["ID", "Nombre Empresa", "NIT/Cédula", "Teléfono", "Dirección", "Correo", "Contacto"]),
    ("Movimientos de Inventario", ["ID", "Producto", "Proveedor", "Precio", "Precio Balance", "Cantidad", "Tipo Movimiento", "Fecha", "Motivo"]),
    ("Facturas", ["ID", "N. Factura", "Cliente", "Fecha", "Método Pago", "Subtotal", "Impuestos", "Total", "Estado"]),
    ("Usuarios", ["ID", "Nombre", "Email", "Rol"]),
])

FILTROS_POR_TIPO = {
    "Productos": ["Todos", "Activo", "Inactivo"],
    "Movimientos de Inventario": ["Todos", "Entrada", "Salida", "Ajuste"],
    "Facturas": ["Todos", "Pagada", "Pendiente", "Anulada"],
}

TIPOS_CON_FECHA = ["Movimientos de Inventario", "Facturas"]


class PanelReportes(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=COLOR_BG, corner_radius=0)
        self.datos_completos = []
        self.datos_actuales = []
        self.columnas_actuales = []
        self.columnas_seleccionadas = {k: list(v) for k, v in COLUMNAS_POR_TIPO.items()}

        # Header
        ctk.CTkLabel(self, text="Generador de Reportes", font=("Segoe UI", 24, "bold"), text_color=COLOR_TEXT, image=icon_chart_pie(24, COLOR_TEXT), compound="left").pack(anchor="w", padx=25, pady=(25, 15))

        # Layout: config panel (left) + table (right)
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=25, pady=(0, 25))

        # --- Left config panel ---
        config_panel = ctk.CTkFrame(body, fg_color=COLOR_CARD, corner_radius=8, width=300)
        config_panel.pack(side="left", fill="y")
        config_panel.pack_propagate(False)

        # Tipo de reporte
        ctk.CTkLabel(config_panel, text="Tipo de Reporte", font=("Segoe UI", 13, "bold"), text_color=COLOR_TEXT_MUTED).pack(anchor="w", padx=15, pady=(15, 5))
        self.combo_tipo = ctk.CTkComboBox(config_panel, values=list(COLUMNAS_POR_TIPO.keys()), width=270,
                                          fg_color="#2d2d2d", button_color=COLOR_BORDER, text_color=COLOR_TEXT,
                                          command=lambda _: self.actualizar_columnas())
        self.combo_tipo.pack(anchor="w", padx=15, pady=(0, 10))
        self.combo_tipo.set(list(COLUMNAS_POR_TIPO.keys())[0])

        # Columnas - scrollable
        ctk.CTkLabel(config_panel, text="Columnas a incluir", font=("Segoe UI", 13, "bold"), text_color=COLOR_TEXT_MUTED).pack(anchor="w", padx=15, pady=(10, 5))

        self.scroll_config = ctk.CTkScrollableFrame(config_panel, fg_color=COLOR_CARD, height=200)
        self.scroll_config.pack(fill="both", expand=True, padx=15, pady=(0, 5))

        # Filtro específico
        self.frame_filtro_esp = ctk.CTkFrame(config_panel, fg_color="transparent")
        self.frame_filtro_esp.pack(fill="x", padx=15, pady=5)

        # Filtro de fecha
        self.frame_fecha = ctk.CTkFrame(config_panel, fg_color="transparent")
        self.frame_fecha.pack(fill="x", padx=15, pady=5)

        # Botones
        btn_frame = ctk.CTkFrame(config_panel, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=10)

        ctk.CTkButton(btn_frame, text="Generar Reporte", fg_color=COLOR_GREEN, hover_color="#059669",
                      font=("Segoe UI", 13, "bold"), height=36, image=icon_chart_bar(16, "#ffffff"), compound="left", command=self.generar_reporte).pack(fill="x", pady=(0, 5))
        ctk.CTkButton(btn_frame, text="Exportar a Excel", fg_color=COLOR_BLUE, hover_color="#2563eb",
                      font=("Segoe UI", 13, "bold"), height=36, image=icon_download(16, "#ffffff"), compound="left", command=self.exportar_excel).pack(fill="x")

        # --- Right table panel ---
        table_panel = ctk.CTkFrame(body, fg_color=COLOR_CARD, corner_radius=8)
        table_panel.pack(side="left", fill="both", expand=True, padx=(10, 0))

        # Search bar
        search_frame = ctk.CTkFrame(table_panel, fg_color="transparent")
        search_frame.pack(fill="x", padx=10, pady=(10, 5))

        self.search_entry = ctk.CTkEntry(search_frame, width=300, fg_color="#2d2d2d", border_color=COLOR_BORDER,
                                         text_color=COLOR_TEXT, placeholder_text="Filtrar resultados...")
        self.search_entry.pack(side="left", padx=(0, 10))
        self.search_entry.bind("<KeyRelease>", lambda e: self.aplicar_filtro_texto())

        self.lbl_registros = ctk.CTkLabel(search_frame, text="0 registros", font=("Segoe UI", 12), text_color=COLOR_TEXT_DIM)
        self.lbl_registros.pack(side="right", padx=5)

        # Tabla
        self.tabla_frame = ctk.CTkFrame(table_panel, fg_color="#2d2d2d", corner_radius=0)
        self.tabla_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.header_tabla = ctk.CTkFrame(self.tabla_frame, fg_color="#3f3f46", height=36, corner_radius=0)
        self.header_tabla.pack(fill="x")
        self.header_tabla.pack_propagate(False)

        self.scroll_tabla = ctk.CTkScrollableFrame(self.tabla_frame, fg_color="#2d2d2d", corner_radius=0)
        self.scroll_tabla.pack(fill="both", expand=True)

        self._inicializar_tabla_vacia()
        self.actualizar_columnas()

    def _inicializar_tabla_vacia(self):
        for w in self.header_tabla.winfo_children():
            w.destroy()
        ctk.CTkLabel(self.header_tabla, text="Seleccione un tipo y genere un reporte", font=("Segoe UI", 13, "bold"), text_color=COLOR_TEXT).pack(side="left", padx=10, pady=8)
        for w in self.scroll_tabla.winfo_children():
            w.destroy()

    def actualizar_columnas(self):
        tipo = self.combo_tipo.get()

        # Columnas checkboxes
        for w in self.scroll_config.winfo_children():
            w.destroy()
        self.chk_columnas = {}
        for col in COLUMNAS_POR_TIPO[tipo]:
            var = ctk.CTkCheckBox(self.scroll_config, text=col, font=("Segoe UI", 12), text_color=COLOR_TEXT,
                                  fg_color=COLOR_GREEN, hover_color="#059669")
            var.select()
            self.chk_columnas[col] = var
            var.pack(anchor="w", pady=1)

        # Filtro específico
        for w in self.frame_filtro_esp.winfo_children():
            w.destroy()
        if tipo in FILTROS_POR_TIPO:
            ctk.CTkLabel(self.frame_filtro_esp, text="Filtrar por", font=("Segoe UI", 13, "bold"), text_color=COLOR_TEXT_MUTED).pack(anchor="w")
            self.combo_filtro_esp = ctk.CTkComboBox(self.frame_filtro_esp, values=FILTROS_POR_TIPO[tipo], width=250,
                                                    fg_color="#2d2d2d", button_color=COLOR_BORDER, text_color=COLOR_TEXT,
                                                    command=lambda _: self._aplicar_filtros())
            self.combo_filtro_esp.pack(anchor="w", pady=(2, 0))
            self.combo_filtro_esp.set("Todos")
        else:
            self.combo_filtro_esp = None

        # Filtro de fecha
        for w in self.frame_fecha.winfo_children():
            w.destroy()
        if tipo in TIPOS_CON_FECHA:
            ctk.CTkLabel(self.frame_fecha, text="Filtrar por fecha", font=("Segoe UI", 13, "bold"), text_color=COLOR_TEXT_MUTED).pack(anchor="w")
            self.chk_fecha = ctk.CTkCheckBox(self.frame_fecha, text="Habilitar filtro de rango", font=("Segoe UI", 12),
                                             text_color=COLOR_TEXT, fg_color=COLOR_GREEN, hover_color="#059669")
            self.chk_fecha.pack(anchor="w", pady=(2, 5))

            fecha_frame = ctk.CTkFrame(self.frame_fecha, fg_color="transparent")
            fecha_frame.pack(fill="x")

            ctk.CTkLabel(fecha_frame, text="Desde:", font=("Segoe UI", 11), text_color=COLOR_TEXT_DIM).pack(anchor="w")
            self.entry_fecha_ini = ctk.CTkEntry(fecha_frame, width=250, fg_color="#2d2d2d", border_color=COLOR_BORDER, text_color=COLOR_TEXT, placeholder_text="YYYY-MM-DD")
            self.entry_fecha_ini.pack(anchor="w", pady=(1, 5))
            self.entry_fecha_ini.configure(state="disabled")

            ctk.CTkLabel(fecha_frame, text="Hasta:", font=("Segoe UI", 11), text_color=COLOR_TEXT_DIM).pack(anchor="w")
            self.entry_fecha_fin = ctk.CTkEntry(fecha_frame, width=250, fg_color="#2d2d2d", border_color=COLOR_BORDER, text_color=COLOR_TEXT, placeholder_text="YYYY-MM-DD")
            self.entry_fecha_fin.pack(anchor="w", pady=(1, 0))
            self.entry_fecha_fin.configure(state="disabled")

            self.chk_fecha.configure(command=self._toggle_fecha)
        else:
            self.chk_fecha = None

    def _toggle_fecha(self):
        if self.chk_fecha.cget("state") == "normal" and hasattr(self, 'chk_fecha_var'):
            pass
        enabled = self.chk_fecha.cget("state") != "disabled"
        # CTkCheckBox doesn't have a simple get; check via variable
        try:
            checked = bool(self.chk_fecha.cget("variable").get())
        except Exception:
            checked = True
        state = "normal" if checked else "disabled"
        self.entry_fecha_ini.configure(state=state)
        self.entry_fecha_fin.configure(state=state)

    def _get_columnas_seleccionadas(self):
        tipo = self.combo_tipo.get()
        return [col for col in COLUMNAS_POR_TIPO[tipo] if self.chk_columnas[col].cget("state") != "disabled" and self._is_checked(self.chk_columnas[col])]

    def _is_checked(self, chk):
        try:
            return bool(chk.cget("variable").get())
        except Exception:
            return True

    def generar_reporte(self):
        tipo = self.combo_tipo.get()
        cols = self._get_columnas_seleccionadas()
        if not cols:
            messagebox.showwarning("Aviso", "Seleccione al menos una columna.")
            return

        self.columnas_actuales = cols
        self.datos_completos = []
        self.datos_actuales = []

        try:
            if tipo == "Productos":
                self._cargar_productos()
            elif tipo == "Categorías":
                self._cargar_categorias()
            elif tipo == "Clientes":
                self._cargar_clientes()
            elif tipo == "Proveedores":
                self._cargar_proveedores()
            elif tipo == "Movimientos de Inventario":
                self._cargar_movimientos()
            elif tipo == "Facturas":
                self._cargar_facturas()
            elif tipo == "Usuarios":
                self._cargar_usuarios()
        except Exception as ex:
            messagebox.showerror("Error", f"Error al cargar datos:\n{ex}")
            return

        self._aplicar_filtros()

    def _aplicar_filtros(self):
        # Filtro específico
        tipo = self.combo_tipo.get()
        if self.combo_filtro_esp and tipo in FILTROS_POR_TIPO:
            valor = self.combo_filtro_esp.get()
            if valor and valor != "Todos":
                self.datos_actuales = [fila for fila in self.datos_completos if any(str(v).lower() == valor.lower() for v in fila)]
            else:
                self.datos_actuales = list(self.datos_completos)
        else:
            self.datos_actuales = list(self.datos_completos)

        self.aplicar_filtro_texto()

    def aplicar_filtro_texto(self):
        texto = self.search_entry.get().strip().lower()
        self._pintar_tabla_header()
        for w in self.scroll_tabla.winfo_children():
            w.destroy()

        count = 0
        for fila in self.datos_actuales:
            if texto:
                match = any(texto in str(v).lower() for v in fila)
                if not match:
                    continue
            row = ctk.CTkFrame(self.scroll_tabla, fg_color="#333333", corner_radius=4)
            row.pack(fill="x", pady=1)
            for v in fila:
                ctk.CTkLabel(row, text=str(v), font=("Segoe UI", 12), text_color=COLOR_TEXT).pack(side="left", padx=10, pady=6)
            count += 1

        self.lbl_registros.configure(text=f"{count} de {len(self.datos_actuales)} registro(s)")

    def _pintar_tabla_header(self):
        for w in self.header_tabla.winfo_children():
            w.destroy()
        for col in self.columnas_actuales:
            ctk.CTkLabel(self.header_tabla, text=col, font=("Segoe UI", 13, "bold"), text_color=COLOR_TEXT).pack(side="left", padx=10, pady=8)

    def _extraer_fila(self, valores_map, cols):
        return [valores_map.get(c, "") for c in cols]

    def _cargar_productos(self):
        dao = ProductoDAO()
        for p in dao.listar_todos_con_categoria():
            v = OrderedDict([
                ("ID", p.id), ("SKU", p.sku or ""), ("Nombre", p.nombre or ""),
                ("Descripción", p.descripcion or ""), ("Categoría", getattr(p, 'categoria_nombre', '') or ""),
                ("Precio Venta", f"${p.precio_venta:.2f}"), ("Costo Promedio", f"${p.costo_promedio:.2f}"),
                ("Stock Mínimo", f"{p.stock_minimo:.0f}"), ("Stock Actual", f"{p.stock_actual:.0f}"),
                ("Estado", "Activo" if p.state else "Inactivo"),
            ])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _cargar_categorias(self):
        dao = CategoriaDAO()
        for c in dao.listar_todas():
            v = OrderedDict([("ID", c.id), ("Nombre", c.nombre or ""), ("Descripción", c.descripcion or "")])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _cargar_clientes(self):
        dao = ClienteDAO()
        for c in dao.listar_todos():
            v = OrderedDict([("ID", c.id), ("Cédula", c.cedula or ""), ("Nombre", c.nombre or ""),
                             ("Correo", c.correo or ""), ("Teléfono", c.telefono or "")])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _cargar_proveedores(self):
        dao = ProveedorDAO()
        for p in dao.listar_todos():
            v = OrderedDict([("ID", p.id), ("Nombre Empresa", p.nombre_empresa or ""),
                             ("NIT/Cédula", p.nit_cedula or ""), ("Teléfono", p.telefono or ""),
                             ("Dirección", p.direccion or ""), ("Correo", p.correo or ""),
                             ("Contacto", p.nombre_contacto or "")])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _cargar_movimientos(self):
        dao = InventarioDAO()
        lista = dao.listar_movimientos_con_nombres()
        fecha_ini = self._get_fecha_ini()
        fecha_fin = self._get_fecha_fin()
        for m in lista:
            if fecha_ini and m.fecha_movimiento:
                if m.fecha_movimiento.date() < fecha_ini:
                    continue
            if fecha_fin and m.fecha_movimiento:
                if m.fecha_movimiento.date() > fecha_fin:
                    continue
            v = OrderedDict([
                ("ID", m.id), ("Producto", m.producto_nombre or ""), ("Proveedor", m.proveedor_nombre or ""),
                ("Precio", f"${m.precio:.2f}"), ("Precio Balance", f"${m.precio_balance:.2f}"),
                ("Cantidad", f"{m.cantidad:.0f}"), ("Tipo Movimiento", m.tipo_movimiento or ""),
                ("Fecha", m.fecha_movimiento.strftime(SDF_FMT) if m.fecha_movimiento else ""),
                ("Motivo", m.motivo or ""),
            ])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _cargar_facturas(self):
        dao = FacturaDAO()
        lista = dao.listar_todas_con_cliente()
        fecha_ini = self._get_fecha_ini()
        fecha_fin = self._get_fecha_fin()
        for f in lista:
            if fecha_ini and f.fecha_emision:
                if f.fecha_emision.date() < fecha_ini:
                    continue
            if fecha_fin and f.fecha_emision:
                if f.fecha_emision.date() > fecha_fin:
                    continue
            v = OrderedDict([
                ("ID", f.id), ("N. Factura", f.numero_factura or ""), ("Cliente", f.cliente_nombre or ""),
                ("Fecha", f.fecha_emision.strftime(SDF_FMT) if f.fecha_emision else ""),
                ("Método Pago", f.metodo_pago or ""), ("Subtotal", f"${f.subtotal:.2f}"),
                ("Impuestos", f"${f.impuestos:.2f}"), ("Total", f"${f.total:.2f}"),
                ("Estado", f.estado or ""),
            ])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _cargar_usuarios(self):
        dao = UsuarioDAO()
        for u in dao.listar_todos():
            v = OrderedDict([("ID", u.id), ("Nombre", u.nombre or ""), ("Email", u.email or ""), ("Rol", u.rol or "")])
            self.datos_completos.append(self._extraer_fila(v, self.columnas_actuales))

    def _get_fecha_ini(self):
        if not self.chk_fecha:
            return None
        try:
            checked = bool(self.chk_fecha.cget("variable").get())
        except Exception:
            return None
        if not checked:
            return None
        val = self.entry_fecha_ini.get().strip()
        if not val:
            return None
        try:
            return datetime.strptime(val, SDF_FMT).date()
        except ValueError:
            return None

    def _get_fecha_fin(self):
        if not self.chk_fecha:
            return None
        try:
            checked = bool(self.chk_fecha.cget("variable").get())
        except Exception:
            return None
        if not checked:
            return None
        val = self.entry_fecha_fin.get().strip()
        if not val:
            return None
        try:
            return datetime.strptime(val, SDF_FMT).date()
        except ValueError:
            return None

    def _abrir_archivo(self, filename):
        try:
            libre_office = None
            for ruta in [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
                r"C:\Program Files\LibreOffice\program\scalc.exe",
                r"C:\Program Files (x86)\LibreOffice\program\scalc.exe"
            ]:
                if os.path.exists(ruta):
                    libre_office = ruta
                    break
            if libre_office:
                subprocess.Popen([libre_office, filename])
            elif os.name == "nt":
                os.startfile(filename)
            else:
                subprocess.call(["xdg-open", filename])
        except Exception:
            pass

    def exportar_excel(self):
        if not self.columnas_actuales:
            messagebox.showwarning("Aviso", "Genere un reporte primero.")
            return

        tipo = self.combo_tipo.get()
        filename = filedialog.asksaveasfilename(
            title="Guardar Reporte",
            defaultextension=".xlsx",
            initialfile=f"Reporte_{tipo.replace(' ', '_')}.xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = tipo[:31]

            # Header
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = Font(bold=True)
            for c, col in enumerate(self.columnas_actuales, 1):
                cell = ws.cell(row=1, column=c, value=col)
                cell.fill = header_fill
                cell.font = header_font

            # Data - read from visible table (filtered)
            texto = self.search_entry.get().strip().lower()
            r = 2
            for fila in self.datos_actuales:
                if texto:
                    if not any(texto in str(v).lower() for v in fila):
                        continue
                for c, val in enumerate(fila, 1):
                    sval = str(val) if val is not None else ""
                    if sval.startswith("$"):
                        try:
                            ws.cell(row=r, column=c, value=float(sval.replace("$", "").replace(",", "")))
                        except ValueError:
                            ws.cell(row=r, column=c, value=sval)
                    else:
                        try:
                            ws.cell(row=r, column=c, value=int(sval))
                        except ValueError:
                            try:
                                ws.cell(row=r, column=c, value=float(sval))
                            except ValueError:
                                ws.cell(row=r, column=c, value=sval)
                r += 1

            # Auto-size columns
            for c in range(1, len(self.columnas_actuales) + 1):
                ws.column_dimensions[chr(64 + c) if c <= 26 else "A"].width = 20

            wb.save(filename)
            messagebox.showinfo("Éxito", f"Reporte exportado con éxito a:\n{filename}")
            self._abrir_archivo(filename)
        except Exception as ex:
            messagebox.showerror("Error", f"Error al exportar:\n{ex}")
